import os
from openpyxl import load_workbook
import pandas as pd

# Folder containing subfolders with Excel files
base_folder_path = r'D:\\Stock\\AllShares'

# Placeholder for the sector value to be added
#sector_value = ""  # Replace with the desired value

# Function to delete the row containing "nepsealpha_export_price"
def delete_row_with_text(file_path, text):
    wb = load_workbook(file_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # Iterate through each row to find the text
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and text in str(cell.value):
                    ws.delete_rows(cell.row, 1)
                    break
    # Save the workbook after deleting the row
    wb.save(file_path)

# List to hold data from each file
dataframes = []

# Walk through all subfolders and files
for root, dirs, files in os.walk(base_folder_path):
    for file in files:
        if file.endswith('.xlsx'):
            # Full file path
            file_path = os.path.join(root, file)

            # Step 1: Delete the row containing "nepsealpha_export_price"
            delete_row_with_text(file_path, "nepsealpha_export_price")

            # Step 2: Read the modified Excel file and append it to the list
            df = pd.read_excel(file_path)

            # Step 3: Add the 'Sector' column with the specified value
            #df['Sector'] = sector_value

            # Append the DataFrame to the list
            dataframes.append(df)

# Step 4: Concatenate all dataframes into one
merged_df = pd.concat(dataframes, ignore_index=True)

# Step 5: Save the merged data to a new Excel file
output_file = os.path.join(base_folder_path, 'Stock.xlsx')
merged_df.to_excel(output_file, index=False)

print(f"All Excel files have been processed, and the merged data with the 'Sector' column has been saved into {output_file}")
